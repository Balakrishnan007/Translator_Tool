# -*- coding: utf-8 -*-
"""Loads the glossary Excel file into a plain list of dicts."""

import openpyxl

GLOSSARY_PATH = r"D:\Rotpunküchen\dataset\00_Glossary_Rotpunkt.xlsx"


def load_glossary(path: str = GLOSSARY_PATH) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Glossary"]
    headers = [c.value for c in ws[1]]
    terms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(headers, row))
        terms.append(entry)
    return terms


if __name__ == "__main__":
    terms = load_glossary()
    print(f"Loaded {len(terms)} glossary terms\n")
    for t in terms[:5]:
        print(t)

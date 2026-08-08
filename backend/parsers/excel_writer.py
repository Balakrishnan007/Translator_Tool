# -*- coding: utf-8 -*-
"""Writes translated segments back out as a real .xlsx file (spec section 10).

Only ever receives "row"-type segments. Per the export compatibility rule,
Word/PDF sources (mixed prose + tables) can't export to Excel at all, only
a purely tabular Excel source can. So there's no paragraph/table_cell case
to handle here, unlike word_writer.py and pdf_writer.py.
"""

import openpyxl


def write_xlsx(segments: list[dict], output_path: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = {}

    segments = sorted(segments, key=lambda s: s["order"])

    for seg in segments:
        if seg.get("type") != "row":
            continue
        sheet_name = seg.get("sheet") or "Sheet1"
        if sheet_name not in sheets:
            sheets[sheet_name] = wb.create_sheet(sheet_name)
        ws = sheets[sheet_name]

        text = seg.get("translation") or f"[TRANSLATION FAILED: {seg.get('error', 'unknown error')}]"
        original_cells = seg.get("cells") or []
        parts = [p.strip() for p in text.split(" | ")]

        row_idx = seg.get("row_index", ws.max_row + 1)
        if original_cells and len(parts) == len(original_cells):
            for cell_info, value in zip(original_cells, parts):
                ws.cell(row=row_idx, column=cell_info["col_index"], value=value)
        else:
            # Translation didn't preserve the "|" boundaries. Don't
            # silently misalign columns, keep the whole row as one cell.
            ws.cell(row=row_idx, column=1, value=text)

        if seg.get("is_header"):
            for cell in ws[row_idx]:
                cell.font = openpyxl.styles.Font(bold=True)

    if not sheets:
        wb.create_sheet("Sheet1")  # never save a workbook with zero sheets

    wb.save(output_path)
    return output_path
